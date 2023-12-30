import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry

class VoterRegistrationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Voter Registration")
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.configure_style()
        self.create_gui()

    def configure_style(self):
        # Configure ttk.Style for red, black, and white colors
        self.style.configure("TLabel", foreground="white", background="black")
        self.style.configure("TEntry", fieldbackground="white")
        self.style.configure("TButton", foreground="white", background="red", padding=(10, 5))
        self.style.configure("TCheckbutton", foreground="white", background="black")

    def create_gui(self):
        # Make the GUI full screen
        self.root.attributes('-fullscreen', False)

        # Ward Entry
        self.ward_entry = self.create_label_entry("Ward:", row=0, column=0, entry_column=1)

        # Election Option Menu
        self.election_combobox = self.create_label_combobox("Election:", ["MP", "MC", "MCP"], row=0, column=2, combobox_column=3, default_value="MP")

        # MP Booth Entry
        self.mp_booth_entry = self.create_label_entry("MP Booth:", row=0, column=4, entry_column=5)

        # MC Booth Entry
        self.mc_booth_entry = self.create_label_entry("MC Booth:", row=0, column=6, entry_column=7)

        # MLA Booth Entry
        self.mla_booth_entry = self.create_label_entry("MLA Booth:", row=0, column=8, entry_column=9)

        # Voter ID
        self.voter_id_entry = self.create_label_entry("Voter ID:", row=1, column=0, entry_column=1)

        # Serial No
        self.serial_no_entry = self.create_label_entry("Serial No:", row=1, column=2, entry_column=3)

        # Page No
        self.page_no_entry = self.create_label_entry("Page No:", row=1, column=4, entry_column=5)

        # Gender Combobox
        self.gender_combobox = self.create_label_combobox("Gender:", ["Male", "Female", "Others"], row=2, column=0, combobox_column=1, default_value="Male")

        # Age
        self.age_entry = self.create_label_entry("Age:", row=2, column=2, entry_column=3)

        # Birthdate
        self.birthdate_label = ttk.Label(self.root, text="Birthdate:")
        self.birthdate_label.grid(row=2, column=4, sticky=tk.W, padx=5, pady=5)
        self.birthdate_entry = ttk.Entry(self.root, style="TEntry", state='readonly')
        self.birthdate_entry.grid(row=2, column=5, padx=5, pady=5)

        # Phone Num
        self.phone_num_entry = self.create_label_entry("Phone Num:", row=2, column=6, entry_column=7)

        # Name
        self.name_entry = self.create_label_entry("Name:", row=3, column=0, entry_column=1)

        # Father/Husband Name
        self.father_husband_name_entry = self.create_label_entry("Father/Husband Name:", row=3, column=2, entry_column=3)

        # Address
        self.address_entry = self.create_label_entry("Address:", row=4, column=0, entry_column=1, column_span=9, width=200)

        # Family
        self.family_entry = self.create_label_entry("Family:", row=5, column=0, entry_column=1)

        # Caste
        self.caste_entry = self.create_label_entry("Caste:", row=5, column=2, entry_column=3)

        # Area Guide
        self.area_guide_entry = self.create_label_entry("Area Guide:", row=5, column=4, entry_column=5)

        # Guide Number
        self.guide_number_entry = self.create_label_entry("Guide Number:", row=5, column=6, entry_column=7)

        # Event Section
        self.event_checkbox_var = tk.BooleanVar()
        self.event_checkbox = ttk.Checkbutton(self.root, text="Event", variable=self.event_checkbox_var,
                                              command=self.toggle_event_fields, style="TCheckbutton")
        self.event_checkbox.grid(row=6, column=0, columnspan=2, pady=5)

        self.date_label = ttk.Label(self.root, text="Date:")
        self.date_entry = DateEntry(self.root, state='disabled', style="TEntry")
        self.date_label.grid(row=6, column=2, sticky=tk.W, padx=5, pady=5)
        self.date_entry.grid(row=6, column=3, padx=5, pady=5)

        self.event_label = ttk.Label(self.root, text="Event:")
        self.event_entry = ttk.Entry(self.root, state='disabled', style="TEntry")
        self.event_label.grid(row=6, column=4, sticky=tk.W, padx=5, pady=5)
        self.event_entry.grid(row=6, column=5, padx=5, pady=5)

        # Memo
        self.memo_entry = self.create_label_entry("Memo:", row=7, column=0, entry_column=1, column_span=9, width=200)

        # Qualification
        self.qualification_entry = self.create_label_entry("Qualification:", row=8, column=0, entry_column=1, column_span=1)

        # Stay
        self.stay_entry = self.create_label_entry("Stay:", row=9, column=0, entry_column=1, column_span=1)

        # Voting Place
        self.voting_place_entry = self.create_label_entry("Voting Place:", row=10, column=0, entry_column=1)

        # Add Party
        self.party_entry = self.create_label_entry("Party:", row=10, column=2, entry_column=3)

        # Add Party ID
        self.party_id_entry = self.create_label_entry("Party ID:", row=10, column=4, entry_column=5)

        # Voted
        self.voted_entry = self.create_label_entry("Voted:", row=11, column=0, entry_column=1)

        # Voted Date
        self.voted_date_entry = self.create_label_entry("Voted Date:", row=11, column=2, entry_column=3)

        # Add Button
        self.add_button = ttk.Button(self.root, text="Add", command=self.save_data, style="TButton")
        self.add_button.grid(row=12, column=0, columnspan=2, pady=10)

        # Next Button
        self.next_button = ttk.Button(self.root, text="Next", command=self.clear_entry_fields, style="TButton")
        self.next_button.grid(row=12, column=2, columnspan=2, pady=10)

        # Font Size Dropdown Menu
        self.font_size_var = tk.StringVar(self.root)
        self.font_size_var.set("Font Size")
        self.font_size_menu = tk.OptionMenu(self.root, self.font_size_var, "8", "10", "12", "14", "16", "18", "20", command=self.change_font_size)
        self.font_size_menu.config(width=10)
        self.font_size_menu.grid(row=12, column=4, padx=5, pady=5)

    def create_label_entry(self, text, row, column, entry_column, column_span=1, width=None):
        label = ttk.Label(self.root, text=text, style="TLabel")
        label.grid(row=row, column=column, sticky=tk.W, padx=5, pady=5, columnspan=column_span)
        entry = ttk.Entry(self.root, style="TEntry", width=width)
        entry.grid(row=row, column=entry_column, padx=5, pady=5, columnspan=column_span)
        return entry

    def create_label_combobox(self, text, values, row, column, combobox_column, default_value=None):
        label = ttk.Label(self.root, text=text, style="TLabel")
        label.grid(row=row, column=column, sticky=tk.W, padx=5, pady=5)
        combobox_var = tk.StringVar()
        combobox = ttk.Combobox(self.root, textvariable=combobox_var, values=values, style="TEntry")
        combobox.grid(row=row, column=combobox_column, padx=5, pady=5)
        if default_value:
            combobox.set(default_value)
        return combobox

    def toggle_event_fields(self):
        if self.event_checkbox_var.get():
            self.date_entry.config(state='normal')
            self.event_entry.config(state='normal')
        else:
            self.date_entry.config(state='disabled')
            self.event_entry.config(state='disabled')

    def save_data(self):
        ward_name = self.ward_entry.get()

        # Create or load the Excel workbook and sheet
        try:
            workbook = load_workbook(f"{ward_name}.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = ward_name

            # Write headings to Excel sheet
            headings = ["Ward", "Election", "MP Booth", "MC Booth", "MLA Booth", "Voter ID", "Serial No", "Page No", "Gender",
                        "Age", "Birthdate", "Phone Num", "Name", "Father/Husband Name", "Address",
                        "Family", "Caste", "Area Guide", "Guide Number", "Event", "Memo", "Qualification",
                        "Stay", "Voting Place", "Party", "Party ID", "Voted", "Voted Date"]
            sheet.append(headings)

        # Get the next available row in the sheet
        next_row = sheet.max_row + 1

        # Get data from entry fields
        data = [
            self.ward_entry.get(),
            self.election_combobox.get(),
            self.mp_booth_entry.get(),
            self.mc_booth_entry.get(),
            self.mla_booth_entry.get(),
            self.voter_id_entry.get(),
            self.serial_no_entry.get(),
            self.page_no_entry.get(),
            self.gender_combobox.get(),
            self.age_entry.get(),
            self.birthdate_entry.get(),
            self.phone_num_entry.get(),
            self.name_entry.get(),
            self.father_husband_name_entry.get(),
            self.address_entry.get(),
            self.family_entry.get(),
            self.caste_entry.get(),
            self.area_guide_entry.get(),
            self.guide_number_entry.get(),
            self.event_checkbox_var.get(),
            self.memo_entry.get(),
            self.qualification_entry.get(),
            self.stay_entry.get(),
            self.voting_place_entry.get(),
            self.party_entry.get(),
            self.party_id_entry.get(),
            self.voted_entry.get(),
            self.voted_date_entry.get()
        ]

        # If the event checkbox is selected, add Date and Event data
        if self.event_checkbox_var.get():
            data.extend([self.date_entry.get(), self.event_entry.get()])
        else:
            # If the event checkbox is not selected, add empty values for Date and Event
            data.extend(["", ""])

        # Write data to Excel sheet
        sheet.append(data)

        # Save the workbook
        workbook.save(f"{ward_name}.xlsx")

    def clear_entry_fields(self):
        for entry in [
            self.ward_entry, self.election_combobox, self.mp_booth_entry, self.mc_booth_entry, self.mla_booth_entry,
            self.voter_id_entry, self.serial_no_entry, self.page_no_entry, self.gender_combobox,
            self.age_entry, self.birthdate_entry, self.phone_num_entry, self.name_entry,
            self.father_husband_name_entry, self.address_entry, self.family_entry,
            self.caste_entry, self.area_guide_entry, self.guide_number_entry, self.memo_entry,
            self.qualification_entry, self.stay_entry, self.voting_place_entry, self.party_entry,
            self.party_id_entry, self.voted_entry, self.voted_date_entry
        ]:
            entry.delete(0, tk.END)

        # Clear date and event entry fields
        self.date_entry.set_date(None)
        self.event_entry.delete(0, tk.END)

        # Uncheck the event checkbox
        self.event_checkbox_var.set(False)

    def change_font_size(self, size):
        # Change font size for all widgets
        new_font = ("TkDefaultFont", size)

        # Set default font for all widgets
        default_font = tk.font.nametofont("TkDefaultFont")
        default_font.configure(size=size)

        # If you are using custom styles, you can update their font size as well
        for style_name in self.style.theme_names():
            self.style.configure(style_name, font=new_font)

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    app = VoterRegistrationApp(root)
    app.run()
